#!/usr/bin/env python3
"""
build_mrc_control_action_matrix.py

Generates execution-plan/MRC-CONTROL-ACTION-RACI-MATRIX.xlsx (+ a matching
.csv export of the same data) by parsing all 110 Master (MRC-###) cards in
mrc-cards/master/ and pulling, per card:

  - The exact card title (Section H1 heading)
  - Its applicable JSIG control(s) (Section 2. References) -- up to 4
    "Control N" columns; cards with more than 4 controls (only MRC-071,
    the annual "-1 policies" review spanning all 17 JSIG families) get a
    single combined descriptive string in Control 1 instead of forcing a
    5th+ column for one outlier row.
  - The actions to accomplish the check -- NOT extracted from the Master
    card's own Section 5 Procedure (which remains generic Stub-status
    template language -- Pattern A-H -- since the Master cards themselves
    are intentionally NOT being hand-upgraded to Guide status). Instead
    this column is populated from a curated, hand-authored sidecar file,
    mrc_master_actions.json (same directory as this script), which maps
    each MRC number to real, control-specific, auditable action steps
    (named tools/commands/consoles/reports, e.g. Active Directory/GPMC/
    PowerShell cmdlets, Trellix/McAfee ePO, Tenable Nessus, Splunk, or a
    realistic organizational-process equivalent for physical/personnel/
    policy-only controls). This is a deliberate, documented divergence:
    the spreadsheet's Actions column is more specific than the Master
    card's own Stub-status Section 5 Procedure. If a Master card is later
    upgraded to Guide status (mirroring the MRC-OPS-### precedent), update
    mrc_master_actions.json to match (or remove its override) so the two
    stay in sync -- there is no automatic reconciliation between them.
  - A RACI responsibility reference (Section 3. Personnel / RACI table)

Scope is intentionally Master-only: the MRC-OPS-### (34 cards) and
MRC-NET-### (16 cards) families carry no JSIG control mapping (each states
"JSIG relationship: none" in its own Section 2), so they are out of scope
for a control/action/RACI matrix by definition.

Read-only against the card markdown files -- never edits mrc-cards/master/
or MAINTENANCE-PLAN.md. Regenerate any time after a master-card content
change (or after editing mrc_master_actions.json) with:

    python3 execution-plan/tools/build_mrc_control_action_matrix.py
"""
import csv
import json
import os
import re
import sys

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
MASTER_DIR = os.path.join(REPO_ROOT, "execution-plan", "mrc-cards", "master")
XLSX_OUT = os.path.join(REPO_ROOT, "execution-plan", "MRC-CONTROL-ACTION-RACI-MATRIX.xlsx")
CSV_OUT = os.path.join(REPO_ROOT, "execution-plan", "MRC-CONTROL-ACTION-RACI-MATRIX.csv")
CURATED_ACTIONS_PATH = os.path.join(os.path.dirname(__file__), "mrc_master_actions.json")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print(
        "ERROR: openpyxl is required (pip install openpyxl). "
        "The .csv export has no such dependency if you only need that file.",
        file=sys.stderr,
    )
    raise

TITLE_RE = re.compile(r"^# (MRC-\d+) — (.+)$")
CONTROL_RE = re.compile(
    r"^- \*\*([A-Z]{2}-\d+(?:\(\d+\))?)\*\*\s+—\s+(.+?)\s+\(\[source\]"
)
RAW_CITATION_RE = re.compile(r"^- Raw calendar citation:\s*(.+)$")
PATTERN_HEADER_RE = re.compile(r"^\*\*(Pattern [A-H] — [^*]+)\*\*")

HEADERS = [
    "MRC Number",
    "Card Title",
    "Control 1",
    "Control 2",
    "Control 3",
    "Control 4",
    "Actions to Accomplish the Check",
    "RACI — Responsibility Reference",
]

MAX_SLOTS = 4


def extract_section(text, start_marker, end_marker):
    start = text.find(start_marker)
    if start == -1:
        return ""
    end = text.find(end_marker, start)
    if end == -1:
        end = len(text)
    return text[start:end]


def parse_card(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    lines = text.splitlines()

    # Title
    m = TITLE_RE.match(lines[0])
    mrc_num, title_text = m.group(1), m.group(2)
    card_title = f"{mrc_num} — {title_text}"

    # Controls (Section 2. References)
    ref_section = extract_section(text, "## 2. References", "## 3.")
    controls = []
    for line in ref_section.splitlines():
        cm = CONTROL_RE.match(line.strip())
        if cm:
            controls.append((cm.group(1), cm.group(2)))
    if not controls:
        for line in ref_section.splitlines():
            rm = RAW_CITATION_RE.match(line.strip())
            if rm:
                controls.append(
                    ("N/A", rm.group(1) + " (raw calendar citation — no specific control ID resolved)")
                )

    # RACI (Section 3. Personnel / RACI)
    raci_section = extract_section(text, "## 3. Personnel / RACI", "## 4.")
    raci_lines = [l for l in raci_section.splitlines() if l.strip().startswith("|")]
    responsible = accountable = consulted = informed = ""
    if len(raci_lines) >= 3:
        data_row = raci_lines[2]
        cells = [c.strip() for c in data_row.strip("|").split("|")]
        cells = [c.replace("**", "") for c in cells]
        if len(cells) >= 4:
            responsible, accountable, consulted, informed = cells[0], cells[1], cells[2], cells[3]

    # Procedure (Section 5. Procedure)
    proc_section = extract_section(text, "## 5. Procedure", "## 6.")
    pattern_header = ""
    for line in proc_section.splitlines():
        pm = PATTERN_HEADER_RE.match(line.strip())
        if pm:
            pattern_header = pm.group(1)
            break

    steps = []
    for line in proc_section.splitlines():
        line = line.strip()
        sm = re.match(r"^(\d+)\.\s+(.*)$", line)
        if sm:
            step_text = sm.group(2)
            action = step_text.split(" — Expected result:")[0].strip()
            steps.append(f"{sm.group(1)}. {action}")

    actions_text = pattern_header
    if steps:
        actions_text += "\n" + "\n".join(steps)

    return {
        "mrc_num": mrc_num,
        "card_title": card_title,
        "controls": controls,
        "responsible": responsible,
        "accountable": accountable,
        "consulted": consulted,
        "informed": informed,
        "actions_text": actions_text,
    }


def mrc_sort_key(num_str):
    m = re.search(r"(\d+)", num_str)
    return int(m.group(1)) if m else 0


def build_row(card):
    """Returns the flat list of 8 cell values for one card, shared by both
    the XLSX and CSV writers so the two outputs never drift apart."""
    controls = card["controls"]
    slot_texts = ["", "", "", ""]
    if len(controls) <= MAX_SLOTS:
        for i, (cid, ctitle) in enumerate(controls):
            slot_texts[i] = f"{cid} — {ctitle}"
    else:
        # Overflow case (only MRC-071, 17 controls): combine the full list
        # into Control 1 as one descriptive string; leave slots 2-4 blank.
        ids = ", ".join(cid for cid, _ in controls)
        slot_texts[0] = (
            f"{len(controls)} controls (one per JSIG family, each that family's "
            f"POLICY AND PROCEDURES control): {ids}"
        )

    raci_text = (
        f"Responsible: {card['responsible']}\n"
        f"Accountable: {card['accountable']}\n"
        f"Consulted: {card['consulted']}\n"
        f"Informed: {card['informed']}"
    )

    return [card["mrc_num"], card["card_title"]] + slot_texts + [card["actions_text"], raci_text]


def load_curated_actions():
    with open(CURATED_ACTIONS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_cards():
    files = sorted(
        (p for p in os.listdir(MASTER_DIR) if p.startswith("MRC-") and p.endswith(".md")),
        key=mrc_sort_key,
    )
    cards = [parse_card(os.path.join(MASTER_DIR, p)) for p in files]

    curated_actions = load_curated_actions()
    missing_curated = []
    for c in cards:
        curated = curated_actions.get(c["mrc_num"])
        if curated:
            c["actions_text"] = curated
        else:
            missing_curated.append(c["mrc_num"])
    if missing_curated:
        print(
            "WARNING - no curated action text found, falling back to Stub "
            "template language for:", missing_curated,
        )

    no_control = [c["mrc_num"] for c in cards if not c["controls"]]
    if no_control:
        print("WARNING - cards with zero controls parsed:", no_control)

    dist = {}
    for c in cards:
        n = len(c["controls"])
        dist[n] = dist.get(n, 0) + 1
    print(f"Parsed {len(cards)} master cards. Control-count distribution:", sorted(dist.items()))

    return cards


def write_xlsx(cards, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MRC Master Control-Action-RACI"

    ACCENT = "01696F"
    header_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D4D1CA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = border

    row_idx = 2
    for c in cards:
        row_vals = build_row(c)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
            cell.border = border
            if col_idx == 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        row_idx += 1

    widths = {"A": 12, "B": 46, "C": 34, "D": 34, "E": 34, "F": 34, "G": 60, "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.auto_filter.ref = f"A1:H{row_idx - 1}"

    wb.save(out_path)
    print(f"Wrote {out_path} ({row_idx - 2} data rows)")


def write_csv(cards, out_path):
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for c in cards:
            writer.writerow(build_row(c))
    print(f"Wrote {out_path} ({len(cards)} data rows)")


def main():
    cards = load_cards()
    write_xlsx(cards, XLSX_OUT)
    write_csv(cards, CSV_OUT)


if __name__ == "__main__":
    main()
